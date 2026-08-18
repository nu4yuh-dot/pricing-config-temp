#!/usr/bin/env python3
"""
Generate golden fixtures for the directional franchise rate card.

Same method, and same reason, as scripts/generate_fixtures.py: the workbook ships with
no cached formula values, so the expected numbers do not exist until a calc engine
produces them. This drives the workbook's own `Calculator` sheet -- the only
service-aware formula chain in the file -- and captures what it computes.

Duplicate `Calculator` once per case, write the inputs, recalculate the whole workbook
in one headless LibreOffice pass, read every duplicate back.

Output: src/pricing/__fixtures__/franchise-golden.json

Usage: python3 scripts/generate_franchise_fixtures.py [--workbook-dir DIR]
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from dataclasses import dataclass, asdict
from pathlib import Path

import openpyxl

SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

REPO = Path(__file__).resolve().parent.parent
WORK = REPO / "scripts" / ".recalc-franchise"
OUT = REPO / "src" / "pricing" / "__fixtures__" / "franchise-golden.json"

WORKBOOK = "DNS_Directional_RateCard_Calculator.xlsx"

# Calculator inputs (the yellow cells).
IN_PINCODE, IN_WEIGHT, IN_SERVICE, IN_VALUE = "C5", "C6", "C7", "C8"
IN_FUEL_AIR, IN_FUEL_SURFACE = "C9", "C10"

# Everything the sheet derives, captured as the expected result.
OUT_CELLS = {
    "zone": "C11",
    "odaStatus": "C12",
    "edlKm": "C13",
    "destination": "C14",
    "chargeableWeight": "C15",
    "freight": "C17",
    "oda": "C18",
    "fuel": "C19",
    "awb": "C20",
    "fov": "C21",
    "subTotal": "C22",
    "gst": "C23",
    "total": "C24",
}


@dataclass
class Case:
    id: str
    description: str
    pincode: int
    weight: float
    service: str
    declaredValue: float = 5000
    fuelAir: float = 0.92
    fuelSurface: float = 0.65


def index_pincodes(wb) -> dict:
    """One representative pincode per zone, plus ODA examples of each kind."""
    ws = wb["Pincode Master"]
    plain: dict[str, int] = {}
    oda_banded: list[tuple[int, str, int]] = []
    oda_far: list[tuple[int, str, int]] = []
    below_range: list[int] = []
    not_in_apex: list[int] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        pin, _area, _district, _state, zone, status, km = row[:7]
        if pin is None:
            continue
        km = km or 0
        if status == "Non-ODA" and zone not in plain:
            plain[zone] = pin
        if str(status).startswith("ODA") and km <= 500 and len(oda_banded) < 4:
            oda_banded.append((pin, zone, km))
        if km > 500 and len(oda_far) < 2:
            oda_far.append((pin, zone, km))
        if status == "Below Range (<20)" and len(below_range) < 1:
            below_range.append(pin)
        if status == "Not in APEX" and len(not_in_apex) < 1:
            not_in_apex.append(pin)

    return {
        "plain": plain,
        "odaBanded": oda_banded,
        "odaFar": oda_far,
        "belowRange": below_range,
        "notInApex": not_in_apex,
    }


def build_cases(idx: dict) -> list[Case]:
    """
    Chosen to pin down every place an off-by-one can hide: each slab edge on both
    incremental services, the per-500g rounding on both document services, both
    minimum-charge floors, every zone, both ODA paths and the FOV minimum.
    """
    plain = idx["plain"]
    cases: list[Case] = []

    west = plain["WEST"]

    # APEX slab edges. The first block covers 5 kg; slabs break at 25, 50 and 100.
    for w in [0.5, 1, 4.9, 5, 5.1, 24.9, 25, 25.1, 49.9, 50, 50.1, 99.9, 100, 100.1, 250, 1000]:
        cases.append(Case(f"apex-west-{w}kg", f"APEX to WEST at {w} kg", west, w, "APEX"))

    # SURFACE slab edges. The first block covers 10 kg; same upper breaks.
    for w in [0.5, 5, 9.9, 10, 10.1, 24.9, 25, 25.1, 49.9, 50, 50.1, 99.9, 100, 100.1, 250, 1000]:
        cases.append(Case(f"surface-west-{w}kg", f"SURFACE to WEST at {w} kg", west, w, "SURFACE"))

    # DOCs: billed per 500 g, rounded up, with a Rs 50 floor and a 0.5 kg minimum.
    for w in [0.1, 0.5, 0.51, 0.9, 1, 1.1, 1.5, 2, 4.9, 5, 6]:
        cases.append(Case(f"docs-west-{w}kg", f"DOCs to WEST at {w} kg", west, w, "DOCs"))

    # DUTS: same per-500 g billing, a Rs 200 floor and a 1 kg minimum.
    for w in [0.5, 1, 1.1, 2, 3, 4.9, 5, 6]:
        cases.append(Case(f"duts-west-{w}kg", f"DUTS to WEST at {w} kg", west, w, "DUTS"))

    # Every zone on every service, at a weight above the first block.
    for zone, pin in plain.items():
        slug = zone.lower().replace(" & ", "-").replace(" ", "-")
        for service in ["DOCs", "DUTS", "APEX", "SURFACE"]:
            weight = 2 if service in ("DOCs", "DUTS") else 30
            cases.append(
                Case(f"{service.lower()}-{slug}-{weight}kg", f"{service} to {zone} at {weight} kg",
                     pin, weight, service)
            )

    # ODA, banded: the matrix is distance x weight, so both axes need exercising.
    for i, (pin, zone, km) in enumerate(idx["odaBanded"]):
        for w in [30, 100, 101, 250, 251, 500, 501, 1001]:
            cases.append(
                Case(f"surface-oda{i}-{w}kg", f"SURFACE to ODA {zone} ({km} km) at {w} kg",
                     pin, w, "SURFACE")
            )
        cases.append(
            Case(f"apex-oda{i}-30kg", f"APEX to ODA {zone} ({km} km) at 30 kg", pin, 30, "APEX")
        )
        # Documents never carry ODA, whatever the destination.
        cases.append(
            Case(f"docs-oda{i}-2kg", f"DOCs to ODA {zone} -- no ODA is charged", pin, 2, "DOCs")
        )

    # Beyond 500 km the matrix stops and a per-km rate takes over.
    for i, (pin, zone, km) in enumerate(idx["odaFar"]):
        for w in [30, 300]:
            cases.append(
                Case(f"surface-far-oda{i}-{w}kg",
                     f"SURFACE to {zone} at {km} km -- the per-km path", pin, w, "SURFACE")
            )

    # Statuses that read like ODA but are not.
    for pin in idx["belowRange"]:
        cases.append(Case("surface-below-range", "SURFACE to a 'Below Range (<20)' pincode",
                          pin, 30, "SURFACE"))
    for pin in idx["notInApex"]:
        cases.append(Case("apex-not-in-apex", "APEX to a 'Not in APEX' pincode", pin, 30, "APEX"))

    # FOV is 0.33% of declared value with a Rs 200 floor: 60,606.06 is where they cross.
    for value in [0, 1000, 60606, 60607, 100000, 1000000]:
        cases.append(Case(f"surface-fov-{value}", f"SURFACE 30 kg, declared value Rs {value}",
                          west, 30, "SURFACE", declaredValue=value))

    # The fuel percentages are advised monthly, so they must not be baked in.
    cases.append(Case("surface-fuel-zero", "SURFACE 30 kg with the fuel surcharge at zero",
                      west, 30, "SURFACE", fuelSurface=0))
    cases.append(Case("apex-fuel-changed", "APEX 30 kg with air fuel at 50%",
                      west, 30, "APEX", fuelAir=0.5))
    cases.append(Case("docs-fuel-changed", "DOCs 2 kg with air fuel at 50% -- documents use air fuel",
                      west, 2, "DOCs", fuelAir=0.5))

    # An unknown pincode must not silently price.
    cases.append(Case("unknown-pincode", "A pincode that is not in the master", 999999, 30, "SURFACE"))

    return cases


def stage_workbook(src: Path, cases: list[Case]) -> Path:
    WORK.mkdir(parents=True, exist_ok=True)
    staged = WORK / src.name
    shutil.copy(src, staged)

    wb = openpyxl.load_workbook(staged)
    base = wb["Calculator"]

    for i, case in enumerate(cases):
        sheet = wb.copy_worksheet(base)
        sheet.title = f"C{i}"
        sheet[IN_PINCODE] = case.pincode
        sheet[IN_WEIGHT] = case.weight
        sheet[IN_SERVICE] = case.service
        sheet[IN_VALUE] = case.declaredValue
        sheet[IN_FUEL_AIR] = case.fuelAir
        sheet[IN_FUEL_SURFACE] = case.fuelSurface

    wb.save(staged)
    return staged


def recalculate(staged: Path) -> Path:
    out_dir = WORK / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(out_dir), str(staged)],
        capture_output=True, text=True,
    )
    recalculated = out_dir / staged.name
    if not recalculated.exists():
        raise RuntimeError(
            f"LibreOffice did not produce {recalculated}.\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )
    return recalculated


def harvest(recalculated: Path, cases: list[Case]) -> list[dict]:
    wb = openpyxl.load_workbook(recalculated, data_only=True)
    results = []
    for i, case in enumerate(cases):
        ws = wb[f"C{i}"]
        results.append({
            "case": asdict(case),
            "expected": {name: ws[ref].value for name, ref in OUT_CELLS.items()},
        })
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook-dir", default=str(Path.home() / "Downloads"))
    args = parser.parse_args()

    if not Path(SOFFICE).exists():
        print(f"LibreOffice not found at {SOFFICE}", file=sys.stderr)
        return 1

    src = Path(args.workbook_dir) / WORKBOOK
    if not src.exists():
        print(f"missing workbook: {src}", file=sys.stderr)
        return 1

    print("indexing pincodes...")
    index_wb = openpyxl.load_workbook(src, read_only=True)
    idx = index_pincodes(index_wb)
    index_wb.close()

    cases = build_cases(idx)
    print(f"{len(cases)} cases; staging...")
    staged = stage_workbook(src, cases)

    print("recalculating in LibreOffice (one pass)...")
    recalculated = recalculate(staged)

    print("harvesting...")
    fixtures = {
        "generatedFrom": f"{WORKBOOK}, recalculated by LibreOffice headless",
        "note": "Expected values are what the workbook's own Calculator computes. "
                "Do not hand-edit; regenerate with scripts/generate_franchise_fixtures.py.",
        "cases": harvest(recalculated, cases),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(fixtures, indent=2) + "\n")
    print(f"wrote {OUT} ({len(cases)} cases)")

    shutil.rmtree(WORK, ignore_errors=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
