#!/usr/bin/env python3
"""
Generate golden pricing fixtures from the source Excel workbooks.

The source workbooks ship with no cached formula values (design spec §2.4 defect 7),
so the expected numbers do not exist until a calc engine produces them. This script
drives the workbooks' own `Rate Calculator` sheet -- the only model-aware formula
chain in the file -- and captures what Excel itself computes.

Method: duplicate `Rate Calculator` once per test case inside a copy of the
workbook, write the inputs into each duplicate, then recalculate the whole workbook
in a single headless LibreOffice pass and read every duplicate back. One recalc per
model instead of one per case.

Output: src/pricing/__fixtures__/golden.json

Usage: python3 scripts/generate_fixtures.py [--workbook-dir DIR]
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from dataclasses import dataclass, asdict
from pathlib import Path

import openpyxl

SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

REPO = Path(__file__).resolve().parent.parent
WORK = REPO / "scripts" / ".recalc"
OUT = REPO / "src" / "pricing" / "__fixtures__" / "golden.json"

MODELS = [
    ("model-1", "DNS_Rate_Card_v15.xlsx", "CUMULATIVE_SLABS"),
    ("model-2", "DNS_Rate_Card_Model2.xlsx", "MIN_PLUS_EXCESS"),
    ("model-3", "DNS_Rate_Card_Model3.xlsx", "MAX_MIN_OR_FULL"),
]

# Rate Calculator cell map (inputs in column B).
IN_MODE, IN_FROM, IN_TO, IN_WEIGHT = "B4", "B5", "B6", "B7"
IN_L, IN_B, IN_H, IN_PIECES, IN_RAIL_HEAVY = "B8", "B9", "B10", "B11", "B12"

# Derived cells we capture as expected output.
OUT_CELLS = {
    "originArea": "B14",
    "originZone": "B15",
    "destArea": "B16",
    "destZone": "B17",
    "originEdlKm": "B18",
    "destEdlKm": "B19",
    "volumetricWeight": "B20",
    "chargeableWeight": "B21",
    "transitTime": "B22",
    "minChargeWeight": "B26",
    "minCharge": "D26",
    "tier1Rate": "C27",
    "tier2Rate": "C28",
    "tier3Rate": "C29",
    "freight": "D30",
    "pickup": "D31",
    "pickupOda": "D32",
    "delivery": "D33",
    "deliveryOda": "D34",
    "fuel": "D35",
    "docket": "D36",
    "subTotal": "D37",
    "gst": "D38",
    "total": "D39",
}


@dataclass
class Case:
    id: str
    description: str
    mode: str
    fromPincode: int
    toPincode: int
    actualWeight: float
    length: float = 0
    breadth: float = 0
    height: float = 0
    pieces: int = 1
    singlePackageOver100kg: bool = False


def build_pincode_index(wb) -> dict:
    """Map zone code -> representative pincodes, per mode, plus useful specials."""
    pm = wb["Pincode Master"]
    by_surface_zone: dict[str, list[int]] = {}
    by_air_zone: dict[str, list[int]] = {}
    oda_surface: list[tuple[int, str, int]] = []
    far_oda_surface: list[tuple[int, str, int]] = []

    for row in pm.iter_rows(min_row=2, values_only=True):
        pin, _area, _state = row[0], row[1], row[2]
        air_zone, air_km = row[5], row[6] or 0
        surf_zone, surf_km = row[11], row[12] or 0
        if pin is None:
            continue
        if surf_zone and surf_km == 0:
            by_surface_zone.setdefault(surf_zone, [])
            if len(by_surface_zone[surf_zone]) < 3:
                by_surface_zone[surf_zone].append(pin)
        if air_zone and air_km == 0:
            by_air_zone.setdefault(air_zone, [])
            if len(by_air_zone[air_zone]) < 3:
                by_air_zone[air_zone].append(pin)
        if 0 < surf_km <= 150 and len(oda_surface) < 3:
            oda_surface.append((pin, surf_zone, surf_km))
        if surf_km > 500 and len(far_oda_surface) < 3:
            far_oda_surface.append((pin, surf_zone, surf_km))

    return {
        "surface": by_surface_zone,
        "air": by_air_zone,
        "oda": oda_surface,
        "farOda": far_oda_surface,
    }


def build_cases(idx: dict) -> list[Case]:
    """
    Cases chosen to pin down the places the three models diverge and where
    off-by-one errors hide: every slab boundary, both ODA paths, the volumetric
    rule, the rail heavy-package rule, intra-zone, and unavailable lanes.
    """
    pnq = idx["surface"]["PNQ"][0]
    ncr = idx["surface"]["NCR"][0]
    pnq_air = idx["air"]["PNQ"][0]
    ncr_air = idx["air"]["NCR"][0]
    bom_air = idx["air"]["BOM"][0]
    pnq_second = idx["surface"]["PNQ"][1]
    oda_pin, _oda_zone, _oda_km = idx["oda"][0]
    far_pin, _far_zone, _far_km = idx["farOda"][0]

    cases: list[Case] = []

    # Slab-boundary sweep on one surface lane. Min weight for surface is 50 kg,
    # tiers break at 100 and 300 -- the three methods diverge most across these.
    for w in [1, 25, 49, 50, 51, 99, 100, 101, 200, 299, 300, 301, 500, 1000]:
        cases.append(Case(f"surface-pnq-ncr-{w}kg", f"Surface PNQ->NCR at {w} kg",
                          "Surface", pnq, ncr, w))

    # Same sweep on air, whose min weight is 25 kg rather than 50.
    for w in [1, 24, 25, 26, 99, 100, 101, 300, 301, 1000]:
        cases.append(Case(f"air-pnq-ncr-{w}kg", f"Air PNQ->NCR at {w} kg",
                          "Air", pnq_air, ncr_air, w))

    # Rail: no fuel surcharge, and the heavy-package doubling rule.
    for w in [51, 100, 101, 300, 301]:
        cases.append(Case(f"rail-pnq-ncr-{w}kg", f"Rail PNQ->NCR at {w} kg",
                          "Rail", pnq, ncr, w))
    cases.append(Case("rail-heavy-package", "Rail 150 kg as a single >=100 kg box (2x rule)",
                      "Rail", pnq, ncr, 150, singlePackageOver100kg=True))
    cases.append(Case("rail-heavy-package-below-threshold",
                      "Rail 99 kg flagged as single box -- below the 100 kg threshold",
                      "Rail", pnq, ncr, 99, singlePackageOver100kg=True))

    # NFO is 2x air across all four grids.
    for w in [25, 100, 301]:
        cases.append(Case(f"nfo-pnq-ncr-{w}kg", f"NFO PNQ->NCR at {w} kg",
                          "NFO", pnq_air, ncr_air, w))

    # Lanes across several destinations at a fixed weight.
    for zone in ["BOM", "AMD", "BLR", "CCU", "GAU", "JSR"]:
        dest = idx["surface"][zone][0]
        cases.append(Case(f"surface-pnq-{zone.lower()}-200kg", f"Surface PNQ->{zone} at 200 kg",
                          "Surface", pnq, dest, 200))

    # Intra-zone: pickup and delivery must both be zero.
    cases.append(Case("surface-intra-zone", "Surface within PNQ -- no pickup/delivery",
                      "Surface", pnq, pnq_second, 200))

    # ODA on the destination, both the banded path and the >500 km per-km path.
    for w in [100, 300]:
        cases.append(Case(f"surface-dest-oda-{w}kg", f"Surface to an ODA pincode at {w} kg",
                          "Surface", pnq, oda_pin, w))
    for w in [100, 600]:
        cases.append(Case(f"surface-dest-far-oda-{w}kg",
                          f"Surface to a >500 km ODA pincode at {w} kg (per-km path)",
                          "Surface", pnq, far_pin, w))
    # ODA on the origin exercises the other lookup.
    cases.append(Case("surface-origin-oda", "Surface from an ODA pincode",
                      "Surface", oda_pin, ncr, 200))

    # Volumetric weight wins over actual.
    cases.append(Case("air-volumetric", "Air 10 kg actual, 2 pieces of 100x100x100 cm",
                      "Air", pnq_air, ncr_air, 10, 100, 100, 100, 2))
    cases.append(Case("surface-volumetric", "Surface 10 kg actual, 1 piece of 80x60x50 cm",
                      "Surface", pnq, ncr, 10, 80, 60, 50, 1))

    # Air PNQ->BOM is '-' in the Air Rates matrix: an unavailable lane.
    cases.append(Case("air-unavailable-lane", "Air PNQ->BOM, which the matrix marks '-'",
                      "Air", pnq_air, bom_air, 200))

    # An unknown pincode must not silently price.
    cases.append(Case("unknown-pincode", "Surface from a pincode not in the master",
                      "Surface", 999999, ncr, 200))

    return cases


def stage_workbook(src: Path, cases: list[Case]) -> Path:
    """Copy the workbook and duplicate Rate Calculator once per case."""
    WORK.mkdir(parents=True, exist_ok=True)
    staged = WORK / src.name
    shutil.copy(src, staged)

    wb = openpyxl.load_workbook(staged)
    base = wb["Rate Calculator"]

    for i, case in enumerate(cases):
        sheet = wb.copy_worksheet(base)
        sheet.title = f"C{i}"
        sheet[IN_MODE] = case.mode
        sheet[IN_FROM] = case.fromPincode
        sheet[IN_TO] = case.toPincode
        sheet[IN_WEIGHT] = case.actualWeight
        sheet[IN_L] = case.length
        sheet[IN_B] = case.breadth
        sheet[IN_H] = case.height
        sheet[IN_PIECES] = case.pieces
        sheet[IN_RAIL_HEAVY] = "Yes" if case.singlePackageOver100kg else "No"

    wb.save(staged)
    return staged


def recalculate(staged: Path) -> Path:
    out_dir = WORK / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(out_dir), str(staged)],
        capture_output=True, text=True,
    )
    recalculated = out_dir / staged.name
    if not recalculated.exists():
        raise RuntimeError(
            f"LibreOffice did not produce {recalculated}.\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )
    return recalculated


def harvest(recalculated: Path, cases: list[Case]) -> list[dict]:
    wb = openpyxl.load_workbook(recalculated, data_only=True)
    results = []
    for i, case in enumerate(cases):
        ws = wb[f"C{i}"]
        expected = {name: ws[ref].value for name, ref in OUT_CELLS.items()}
        results.append({"case": asdict(case), "expected": expected})
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook-dir", default=str(Path.home() / "Downloads"))
    args = parser.parse_args()
    src_dir = Path(args.workbook_dir)

    if not Path(SOFFICE).exists():
        print(f"LibreOffice not found at {SOFFICE}", file=sys.stderr)
        return 1

    fixtures: dict = {
        "generatedFrom": "source workbooks recalculated by LibreOffice headless",
        "note": "Expected values are what the workbooks' own Rate Calculator computes. "
                "Do not hand-edit; regenerate with scripts/generate_fixtures.py.",
        "models": [],
    }

    for key, filename, method in MODELS:
        src = src_dir / filename
        if not src.exists():
            print(f"missing workbook: {src}", file=sys.stderr)
            return 1

        print(f"[{key}] indexing pincodes...")
        index_wb = openpyxl.load_workbook(src, read_only=True)
        idx = build_pincode_index(index_wb)
        index_wb.close()

        cases = build_cases(idx)
        print(f"[{key}] staging {len(cases)} cases...")
        staged = stage_workbook(src, cases)
        print(f"[{key}] recalculating...")
        recalculated = recalculate(staged)
        print(f"[{key}] harvesting...")
        results = harvest(recalculated, cases)

        fixtures["models"].append({
            "key": key,
            "sourceFile": filename,
            "freightMethod": method,
            "cases": results,
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(fixtures, indent=2, default=str))
    total = sum(len(m["cases"]) for m in fixtures["models"])
    print(f"\nwrote {OUT.relative_to(REPO)} — {total} cases across {len(fixtures['models'])} models")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
