#!/usr/bin/env python3
"""
Extract the three rate cards and the shared pincode master from the source
workbooks into clean JSON.

This is the boundary between Excel and the application: Python owns reading the
workbooks, TypeScript consumes only the JSON. Nothing downstream needs a
spreadsheet library.

Resolves the source defects catalogued in the design spec §2.4 -- notably it reads
the authoritative `Charges & Terms` column E rather than the stale display copy in
column B, and it locates each parameter by the `(En)` reference written into its
own label rather than by row position.

Output:
  data/extracted/model-1.json, model-2.json, model-3.json
  data/extracted/pincodes.json

Usage: python3 scripts/extract.py [--workbook-dir DIR]
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT_DIR = REPO / "data" / "extracted"

MODELS = [
    ("model-1", "Model 1 — cumulative slabs", "DNS_Rate_Card_v15.xlsx", "CUMULATIVE_SLABS"),
    ("model-2", "Model 2 — minimum + excess weight", "DNS_Rate_Card_Model2.xlsx", "MIN_PLUS_EXCESS"),
    ("model-3", "Model 3 — max of minimum or full weight", "DNS_Rate_Card_Model3.xlsx", "MAX_MIN_OR_FULL"),
]

SURFACE_ZONES = ["PNQ", "PCMC", "KSK", "CSN", "BOM", "NAG", "AMD", "IDR", "NCR", "BWR",
                 "UTR", "LDH", "UPX", "BLR", "HSR", "MAA", "CJB", "HYD", "CCU", "JSR", "GAU"]
AIR_ZONES = ["PNQ", "BOM", "AMD", "IDR", "NCR", "UTR", "BLR", "MAA", "HYD", "CJB", "CCU", "NAG"]

# Each rate sheet stacks four matrices. Values are the header row of each; data
# starts on the row after, and runs for as many rows as the mode has zones.
AIR_GRID_HEADER_ROWS = {"minCharge": 4, "tier1": 19, "tier2": 34, "tier3": 49}
GROUND_GRID_HEADER_ROWS = {"minCharge": 4, "tier1": 28, "tier2": 52, "tier3": 76}


def numeric_or_none(value):
    """Rate cells hold either a number or the workbook's '-' for an unserved lane."""
    if isinstance(value, (int, float)):
        return value
    return None


def read_matrix(ws, header_row: int, zones: list[str]) -> dict:
    """
    Read one origin x destination matrix. Column and row labels are verified
    against the expected zone list rather than assumed, so a reshaped workbook
    fails loudly instead of importing shifted data.
    """
    actual_cols = [ws.cell(row=header_row, column=2 + i).value for i in range(len(zones))]
    if actual_cols != zones:
        raise ValueError(
            f"{ws.title!r} row {header_row}: expected destination columns {zones}, "
            f"got {actual_cols}"
        )

    grid: dict = {}
    for r, origin in enumerate(zones):
        row_index = header_row + 1 + r
        label = ws.cell(row=row_index, column=1).value
        if label != origin:
            raise ValueError(
                f"{ws.title!r} row {row_index}: expected origin {origin!r}, got {label!r}"
            )
        grid[origin] = {
            dest: numeric_or_none(ws.cell(row=row_index, column=2 + c).value)
            for c, dest in enumerate(zones)
        }
    return grid


def read_mode_grids(ws, header_rows: dict, zones: list[str]) -> dict:
    return {name: read_matrix(ws, row, zones) for name, row in header_rows.items()}


def read_charges(ws) -> dict:
    """
    Read the global parameters. Column A labels carry the authoritative cell
    reference, e.g. 'Pickup Surface (E14)' -- we follow that pointer rather than
    trusting column B, which the source lets drift out of sync (defect 3).
    """
    refs: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if not isinstance(label, str):
            continue
        match = re.search(r"\((E\d+)\)", label)
        if match:
            refs[label.split("(")[0].strip()] = match.group(1)

    def param(name: str) -> float:
        ref = refs.get(name)
        if ref is None:
            raise ValueError(f"{ws.title!r}: no cell reference found for {name!r}")
        value = ws[ref].value
        if not isinstance(value, (int, float)):
            raise ValueError(f"{ws.title!r}!{ref} for {name!r} is not numeric: {value!r}")
        return value

    def volumetric_divisor(label_prefix: str) -> float:
        """Stored as a display string such as '/5000' in column B."""
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == label_prefix:
                raw = str(ws.cell(row=row, column=2).value or "")
                digits = re.sub(r"[^0-9]", "", raw)
                if not digits:
                    raise ValueError(f"{ws.title!r}: cannot read divisor from {raw!r}")
                return float(digits)
        raise ValueError(f"{ws.title!r}: no row labelled {label_prefix!r}")

    return {
        "pickupAir": param("Pickup Air"),
        "deliveryAir": param("Delivery Air"),
        "pickupSurface": param("Pickup Surface"),
        "deliverySurface": param("Delivery Surface"),
        "docket": param("Docket"),
        "gstAir": param("GST Air"),
        "gstSurface": param("GST Surface"),
        "minWeightAir": param("Min wt Air"),
        "minWeightSurface": param("Min wt Surface"),
        "volumetricDivisorAir": volumetric_divisor("Volumetric Air"),
        "volumetricDivisorSurface": volumetric_divisor("Volumetric Surface"),
        "fuelAir": param("Fuel Air"),
        "fuelSurface": param("Fuel Surface"),
        # Rail carries no fuel surcharge; the source hardcodes this as a zero
        # multiplier in the calculator rather than storing it as a parameter.
        "fuelRail": 0,
        # Railway parcel norm, stated in Charges & Terms note 6.
        "railHeavyPackageThreshold": 100,
        "railHeavyPackageMultiplier": 2,
        # NFO Rates is every Air Rates cell doubled.
        "nfoMultiplier": 2,
    }


def read_pickup_delivery(ws) -> dict:
    result: dict = {}
    for r, zone in enumerate(SURFACE_ZONES):
        row = 4 + r
        label = ws.cell(row=row, column=1).value
        if label != zone:
            raise ValueError(f"{ws.title!r} row {row}: expected zone {zone!r}, got {label!r}")
        result[zone] = {
            "pickupSurface": ws.cell(row=row, column=4).value,
            "deliverySurface": ws.cell(row=row, column=5).value,
            "pickupAir": ws.cell(row=row, column=6).value,
            "deliveryAir": ws.cell(row=row, column=7).value,
        }
    return result


def read_edl(ws) -> dict:
    km_bands, rates = [], []
    row = 4
    while True:
        km = ws.cell(row=row, column=1).value
        if not isinstance(km, (int, float)):
            break
        km_bands.append(km)
        rates.append([ws.cell(row=row, column=2 + c).value for c in range(5)])
        row += 1
    if not km_bands:
        raise ValueError(f"{ws.title!r}: no EDL km bands found at row 4")
    return {
        "kmBands": km_bands,
        # Column headers are '0-100kg', '101-250', ... -- the lower bound of each.
        "weightBands": [0, 101, 251, 501, 1001],
        "rates": rates,
        "perKmBeyondLastBand": 14,
        "perKmThreshold": 500,
    }


def read_transit(ws, zones: list[str]) -> dict:
    result: dict = {}
    for r, origin in enumerate(zones):
        row = 5 + r
        label = ws.cell(row=row, column=1).value
        if label != origin:
            raise ValueError(f"{ws.title!r} row {row}: expected origin {origin!r}, got {label!r}")
        result[origin] = {
            dest: numeric_or_none(ws.cell(row=row, column=2 + c).value)
            for c, dest in enumerate(zones)
        }
    return result


def read_zone_labels(ws) -> dict:
    """
    Cluster Guide: surface clusters in columns A/B, air hubs in D/E, both from row 3.
    Labels are editable in the dashboard, so they belong in the card data rather
    than hardcoded.
    """
    surface, air = {}, {}
    for row in range(3, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        if isinstance(code, str) and code in SURFACE_ZONES:
            surface[code] = {"belt": ws.cell(row=row, column=2).value or ""}
        air_code = ws.cell(row=row, column=4).value
        if isinstance(air_code, str) and air_code in AIR_ZONES:
            air[air_code] = {"city": ws.cell(row=row, column=5).value or ""}

    missing_surface = set(SURFACE_ZONES) - set(surface)
    missing_air = set(AIR_ZONES) - set(air)
    if missing_surface or missing_air:
        raise ValueError(
            f"{ws.title!r}: missing zone labels — surface {sorted(missing_surface)}, "
            f"air {sorted(missing_air)}"
        )
    return {"surface": surface, "air": air}


def mode_info(row, offset: int, extra_station: bool = False) -> dict:
    """Pincode Master stores three near-identical blocks of per-mode columns."""
    info = {
        "serviceable": row[offset] == "Yes",
        "hub": row[offset + 1],
        "zone": row[offset + 2],
        "edlKm": row[offset + 3] or 0,
        "oda": row[offset + 4] == "Yes",
        "odaCategory": row[offset + 5],
    }
    if extra_station:
        info["station"] = row[offset + 1]
    return info


def read_pincodes(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out.append({
            "pincode": row[0],
            "area": row[1],
            "state": row[2],
            # Columns D..I air, J..O surface, P..U rail.
            "air": mode_info(row, 3),
            "surface": mode_info(row, 9),
            "rail": mode_info(row, 15, extra_station=True),
        })
    return out


def extract_card(path: Path, key: str, name: str, method: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    return {
        "key": key,
        "name": name,
        "freightMethod": method,
        "sourceFile": path.name,
        "data": {
            "grids": {
                "air": read_mode_grids(wb["Air Rates"], AIR_GRID_HEADER_ROWS, AIR_ZONES),
                "surface": read_mode_grids(wb["Surface Rates"], GROUND_GRID_HEADER_ROWS, SURFACE_ZONES),
                "rail": read_mode_grids(wb["Rail Rates"], GROUND_GRID_HEADER_ROWS, SURFACE_ZONES),
            },
            "pickupDelivery": read_pickup_delivery(wb["Pickup & Delivery"]),
            "edlMatrix": read_edl(wb["EDL Matrix"]),
            "transitTimes": {
                "air": read_transit(wb["TAT Air"], AIR_ZONES),
                "surface": read_transit(wb["TAT Surface"], SURFACE_ZONES),
                "rail": read_transit(wb["ETA Rail"], SURFACE_ZONES),
            },
            "charges": read_charges(wb["Charges & Terms"]),
            "zones": read_zone_labels(wb["Cluster Guide"]),
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook-dir", default=str(Path.home() / "Downloads"))
    args = parser.parse_args()
    src_dir = Path(args.workbook_dir)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for key, name, filename, method in MODELS:
        path = src_dir / filename
        if not path.exists():
            print(f"missing workbook: {path}", file=sys.stderr)
            return 1
        card = extract_card(path, key, name, method)
        (OUT_DIR / f"{key}.json").write_text(json.dumps(card, indent=2))
        print(f"wrote {key}.json")

    # The pincode master is identical across all three workbooks (verified by
    # diff), so it is extracted once and shared.
    wb = openpyxl.load_workbook(src_dir / MODELS[0][2], read_only=True)
    pincodes = read_pincodes(wb["Pincode Master"])
    (OUT_DIR / "pincodes.json").write_text(json.dumps(pincodes))
    print(f"wrote pincodes.json — {len(pincodes)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
